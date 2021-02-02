# from app.interactive_imports import *
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import re
from app.utils.features.openpyxl_wrapper import ExcelBlock
import pandas as pd
from flask import current_app
import json


class Cell:
    def __init__(self, row, column, name):
        self.row = row
        self.column = column
        self.name = name


CELLS = {
    'Boiling': Cell(1, 1, 'Тип варки'),
    'FormFactor': Cell(1, 2, 'Форм фактор'),
    'Brand': Cell(1, 3, 'Бренд'),
    'SKU': Cell(1, 4, 'Номенклатура'),
    'FactRemains': Cell(1, 5, 'Факт.остатки, заявка'),
    'NormativeRemains': Cell(1, 6, 'Нормативные остатки'),
    'ProductionPlan': Cell(1, 7, 'План производства'),
    'ExtraPacking': Cell(1, 8, 'Дополнительная фасовка'),
    'Volume': Cell(1, 10, 'Объем варки'),
    'Estimation': Cell(1, 11, 'Расчет'),
    'Plan': Cell(1, 12, 'План'),
    'BoilingVolumes': Cell(1, 13, 'Объемы варок'),
    'Name1': Cell(1, 15, 'Фактические остатки на складах - Заявлено, кг:'),
    'Name2': Cell(2, 15, 'Нормативные остатки, кг')
}


COLUMNS = {
    'BoilingVolume': 10,
    'SKUS_ID': 18,
    'BOILING_ID': 19
}


def generate_title(sheet):
    sheet.column_dimensions[openpyxl.utils.get_column_letter(COLUMNS['SKUS_ID'])].hidden = True
    sheet.column_dimensions[openpyxl.utils.get_column_letter(COLUMNS['BOILING_ID'])].hidden = True
    for key, val in CELLS.items():
        sheet.cell(val.row, val.column).value = val.name
    sheet.freeze_panes = sheet['A2']

    sheet.column_dimensions[get_column_letter(CELLS['Brand'].column)].width = 2 * 5
    sheet.column_dimensions[get_column_letter(CELLS['SKU'].column)].width = 12 * 5
    sheet.column_dimensions[get_column_letter(CELLS['Boiling'].column)].width = 3 * 5


def build_plan_sku(date, df, request_list, plan_path=None):
    template_wb = openpyxl.load_workbook(current_app.config['TEMPLATE_BOILING_PLAN'])
    filename = '{} План по SKU.xlsx'.format(date.strftime('%Y-%m-%d'))
    if plan_path is None:
        path = '{}/{}'.format(current_app.config['SKU_PLAN_FOLDER'], filename)
    else:
        path = plan_path
    template_wb.save(path)
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = template_wb
        writer.sheets = dict((ws.title, ws) for ws in template_wb.worksheets)
        df.to_excel(writer, sheet_name=current_app.config['SHEET_NAMES']['remainings'])
        writer.save()

    wb = openpyxl.load_workbook(filename=path)
    sheet_plan = wb[current_app.config['SHEET_NAMES']['schedule_plan']]

    cur_row, space_rows = 2, 2
    request_list = sorted(request_list, key=lambda k: (k['IsLactose'], k['GroupSKU'][0]['SKU'].made_from_boilings[0].percent))
    is_lactose = False
    for group_skus in request_list:
        if group_skus['GroupSKU'][0]['SKU'].made_from_boilings[0].is_lactose != is_lactose:
            cur_row += space_rows
        is_lactose = group_skus['GroupSKU'][0]['SKU'].made_from_boilings[0].is_lactose
        beg_row = cur_row
        block = ExcelBlock(sheet=sheet_plan, row_height=13.8)
        group_formula = []

        for group in current_app.config['ORDER']:
            block_skus = [x for x in group_skus['GroupSKU'] if x['SKU'].group.name == group]
            block_skus = sorted(block_skus, key=lambda k: k['SKU'].form_factor.relative_weight)
            beg_ff_row = cur_row
            for sku in block_skus:
                formula_plan = "=IFERROR(INDEX('{0}'!$A$5:$DK$265,MATCH($O$1,'{0}'!$A$5:$A$228,0),MATCH({1},'{0}'!$A$5:$DK$5,0)), 0)".format(
                    current_app.config['SHEET_NAMES']['remainings'], block.sheet.cell(cur_row, CELLS['SKU'].column).coordinate)
                formula_remains = "=IFERROR(INDEX('{0}'!$A$5:$DK$265,MATCH($O$2,'{0}'!$A$5:$A$228,0),MATCH({1},'{0}'!$A$5:$DK$5,0)), 0)".format(
                    current_app.config['SHEET_NAMES']['remainings'], block.sheet.cell(cur_row, CELLS['SKU'].column).coordinate)

                block.colour = sku['SKU'].colour[1:]
                block.draw_value(row=cur_row, col=CELLS['Brand'].column, value=sku["SKU"].brand_name)
                block.draw_value(row=cur_row, col=CELLS['SKU'].column, value=sku["SKU"].name)
                block.draw_value(row=cur_row, col=CELLS['FactRemains'].column, value=formula_plan)
                block.draw_value(row=cur_row, col=CELLS['NormativeRemains'].column, value=formula_remains)
                block.draw_value(row=cur_row,
                                 col=CELLS['ProductionPlan'].column,
                                 value='=MIN({}, 0)'.format(block.sheet.cell(cur_row, CELLS['FactRemains'].column).coordinate) if sku['SKU'].production_by_request else 0)
                block.draw_value(row=cur_row,
                                 col=CELLS['ExtraPacking'].column,
                                 value=0 if sku['SKU'].packing_by_request else '=MIN({}, 0)'.format(block.sheet.cell(cur_row, CELLS['FactRemains'].column).coordinate))

                group_formula.append(block.sheet.cell(cur_row, CELLS['ProductionPlan'].column).coordinate)
                cur_row += 1
            end_ff_row = cur_row - 1
            if beg_ff_row <= end_ff_row:
                block.merge_cells(
                    beg_row=beg_ff_row,
                    end_row=end_ff_row,
                    beg_col=CELLS['FormFactor'].column,
                    end_col=CELLS['FormFactor'].column,
                    value=group,
                    alignment=Alignment(horizontal='center', vertical='center', wrapText=True)
                )
        end_row = cur_row - 1
        block.colour = current_app.config['COLOURS']['DefaultGray'][1:]
        block.merge_cells(
            beg_row=beg_row,
            end_row=end_row,
            beg_col=CELLS['Boiling'].column,
            end_col=CELLS['Boiling'].column,
            value='{}% варка, {}{}'.format(
                group_skus["GroupSKU"][0]["SKU"].made_from_boilings[0].percent,
                group_skus["GroupSKU"][0]["SKU"].made_from_boilings[0].ferment,
                ', без лактозы' if not group_skus["GroupSKU"][0]["SKU"].made_from_boilings[0].is_lactose else ''
            ),
            alignment=Alignment(horizontal='center', vertical='center', wrapText=True)
        )
        block.draw_value(row=beg_row, col=CELLS['Volume'].column, value=group_skus['Volume'])
        formula_boiling_count = '{}'.format(str(group_formula).strip('[]').replace(',', ' +').replace('\'', "").upper())
        block.draw_value(row=beg_row,
                         col=CELLS['Estimation'].column,
                         value='=-({}) / {}'.format(
                            formula_boiling_count,
                            block.sheet.cell(beg_row, CELLS['Volume'].column).coordinate.upper()))

        block.draw_value(row=beg_row, col=CELLS['Plan'].column, value='=ROUND({}, 0)'
                         .format(block.sheet.cell(beg_row, CELLS['Estimation'].column).coordinate.upper()))
        block.draw_value(
            row=beg_row,
            col=COLUMNS['SKUS_ID'],
            value=str([x["SKU"].id for x in group_skus["GroupSKU"]])
        )
        block.draw_value(
            row=beg_row,
            col=COLUMNS['BOILING_ID'],
            value=group_skus['BoilingId']
        )
        if is_lactose:
            cur_row += space_rows

    for i, sheet in enumerate(wb.sheetnames):
        if i != 1:
            wb[sheet].views.sheetView[0].tabSelected = False
    wb.active = 1
    wb.save(path)
    return filename


def parse_plan_cell(date, wb, excel, skus):
    sheet_plan = wb[current_app.config['SHEET_NAMES']['remainings']]
    response = {'Date': date, 'WeekDay': date.weekday(), 'Boilings': []}
    for i in range(1, 200):
        if sheet_plan.cell(i, COLUMNS['BOILING_ID']).value is not None:
            boilings_count = excel.evaluate("'{}'!{}".format(
                current_app.config['SHEET_NAMES']['remainings'],
                sheet_plan.cell(i, CELLS['Plan'].column).coordinate)
            )
            sku_ids = sheet_plan.cell(i, COLUMNS['SKUS_ID']).value
            boiling_id = sheet_plan.cell(i, COLUMNS['BOILING_ID']).value
            sku_group = [sku for sku in skus if sku.id in json.loads(sku_ids)]

            sku_volumes = {}
            for j in range(i, 200):
                if sheet_plan.cell(j, CELLS['SKU'].column).value in [x.name for x in sku_group]:
                    sku_id = [x.id for x in sku_group if x.name == sheet_plan.cell(j, CELLS['SKU'].column).value][0]
                    volume = abs(excel.evaluate("'{}'!{}".format(
                        current_app.config['SHEET_NAMES']['remainings'],
                        sheet_plan.cell(j, CELLS['ProductionPlan'].column).coordinate)
                    ))
                    sku_volumes[sku_id] = volume

            boiling_weights = []
            if sheet_plan.cell(i, CELLS['BoilingVolumes'].column).value is not None:
                boiling_weights = re.split(', |. | ', sheet_plan.cell(i, CELLS['BoilingVolumes'].column).value)
                boiling_weights = [x for x in boiling_weights if current_app.config['BOILING_VOLUME_LIMITS']['MIN'] <=
                                   int(x) <= current_app.config['BOILING_VOLUME_LIMITS']['MAX']]
                boiling_weights = [int(x) for x in boiling_weights if isinstance(x, int) or x.isdigit()]

            if len(boiling_weights) > boilings_count:
                boiling_weights = boilings_count * [current_app.config['BOILING_VOLUME_LIMITS']['MAX']]
            else:
                boiling_weights += int(boilings_count - len(boiling_weights)) * [current_app.config['BOILING_VOLUME_LIMITS']['MAX']]
            response['Boilings'].append({
                "BoilingId": boiling_id,
                "BoilingCount": boilings_count,
                "BoilingWeights": boiling_weights,
                "SKUVolumes": sku_volumes
            })
    response['Boilings'] = [x for x in response['Boilings'] if x['BoilingCount'] > 0]
    return response